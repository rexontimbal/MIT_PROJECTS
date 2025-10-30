"""
Export utilities for accidents and clusters
Provides Excel and PDF export functionality
"""

import os
from datetime import datetime
from django.conf import settings
from django.utils import timezone
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import io


class AccidentExporter:
    """
    Excel exporter for accident data
    """

    def __init__(self):
        self.media_root = settings.MEDIA_ROOT
        self.exports_dir = os.path.join(self.media_root, 'exports')
        os.makedirs(self.exports_dir, exist_ok=True)

    def export_to_excel(self, queryset, filename=None):
        """
        Export accidents to Excel with formatting

        Args:
            queryset: Accident queryset
            filename: Optional custom filename

        Returns:
            str: Path to generated Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'accidents_export_{timestamp}.xlsx'

        filepath = os.path.join(self.exports_dir, filename)

        # Convert queryset to dataframe
        data = list(queryset.values(
            'id', 'pro', 'province', 'municipal', 'barangay', 'street',
            'latitude', 'longitude', 'date_committed', 'time_committed',
            'year', 'incident_type', 'offense', 'victim_killed',
            'victim_injured', 'victim_unharmed', 'victim_count',
            'vehicle_kind', 'vehicle_make', 'vehicle_model',
            'is_hotspot', 'cluster_id'
        ))

        df = pd.DataFrame(data)

        # Create workbook with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = 'Accidents Data'

        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write headers
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header.replace('_', ' ').title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add summary sheet
        summary_ws = wb.create_sheet('Summary')
        summary_ws['A1'] = 'Accident Export Summary'
        summary_ws['A1'].font = Font(bold=True, size=14)
        summary_ws['A3'] = 'Total Accidents:'
        summary_ws['B3'] = len(df)
        summary_ws['A4'] = 'Fatal Accidents:'
        summary_ws['B4'] = df['victim_killed'].sum() if 'victim_killed' in df.columns else 0
        summary_ws['A5'] = 'Injury Accidents:'
        summary_ws['B5'] = df['victim_injured'].sum() if 'victim_injured' in df.columns else 0
        summary_ws['A6'] = 'Hotspot Accidents:'
        summary_ws['B6'] = df['is_hotspot'].sum() if 'is_hotspot' in df.columns else 0
        summary_ws['A7'] = 'Export Date:'
        summary_ws['B7'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Save workbook
        wb.save(filepath)

        return filepath

    def export_to_csv(self, queryset, filename=None):
        """
        Export accidents to CSV

        Args:
            queryset: Accident queryset
            filename: Optional custom filename

        Returns:
            str: Path to generated CSV file
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'accidents_export_{timestamp}.csv'

        filepath = os.path.join(self.exports_dir, filename)

        # Convert to dataframe and save
        data = list(queryset.values())
        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False)

        return filepath


class ClusterPDFExporter:
    """
    PDF report generator for accident clusters (hotspots)
    """

    def __init__(self):
        self.media_root = settings.MEDIA_ROOT
        self.exports_dir = os.path.join(self.media_root, 'exports')
        os.makedirs(self.exports_dir, exist_ok=True)

    def generate_report(self, queryset, filename=None):
        """
        Generate PDF report for hotspots

        Args:
            queryset: AccidentCluster queryset
            filename: Optional custom filename

        Returns:
            str: Path to generated PDF file
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'hotspots_report_{timestamp}.pdf'

        filepath = os.path.join(self.exports_dir, filename)

        # Create PDF document
        doc = SimpleDocTemplate(
            filepath,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )

        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        )

        # Title
        title = Paragraph("Accident Hotspot Detection Report", title_style)
        elements.append(title)

        # Report metadata
        report_date = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        metadata = Paragraph(f"<b>Generated:</b> {report_date}", styles['Normal'])
        elements.append(metadata)
        elements.append(Spacer(1, 12))

        # Summary statistics
        total_hotspots = queryset.count()
        total_accidents = sum(cluster.accident_count for cluster in queryset)
        total_casualties = sum(cluster.total_casualties for cluster in queryset)

        summary_heading = Paragraph("Executive Summary", heading_style)
        elements.append(summary_heading)

        summary_data = [
            ['Metric', 'Value'],
            ['Total Hotspots Identified', str(total_hotspots)],
            ['Total Accidents in Hotspots', str(total_accidents)],
            ['Total Casualties', str(total_casualties)],
            ['Average Accidents per Hotspot', f'{total_accidents/max(total_hotspots, 1):.1f}'],
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Top hotspots table
        top_heading = Paragraph("Top 10 Critical Hotspots", heading_style)
        elements.append(top_heading)

        hotspots_data = [['Rank', 'Location', 'Accidents', 'Casualties', 'Severity']]

        for idx, cluster in enumerate(queryset[:10], 1):
            hotspots_data.append([
                str(idx),
                cluster.primary_location[:30],
                str(cluster.accident_count),
                str(cluster.total_casualties),
                f'{cluster.severity_score:.1f}'
            ])

        hotspots_table = Table(hotspots_data, colWidths=[0.6*inch, 2.5*inch, 1*inch, 1*inch, 1*inch])
        hotspots_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))

        elements.append(hotspots_table)
        elements.append(Spacer(1, 20))

        # Detailed hotspot information (each on its own section)
        for cluster in queryset[:5]:  # Top 5 in detail
            detail_heading = Paragraph(
                f"Hotspot Detail: {cluster.primary_location}",
                heading_style
            )
            elements.append(detail_heading)

            detail_data = [
                ['Attribute', 'Value'],
                ['Cluster ID', str(cluster.cluster_id)],
                ['Location', cluster.primary_location],
                ['Coordinates', f'{cluster.center_latitude:.4f}, {cluster.center_longitude:.4f}'],
                ['Total Accidents', str(cluster.accident_count)],
                ['Total Casualties', str(cluster.total_casualties)],
                ['Severity Score', f'{cluster.severity_score:.2f}/100'],
                ['Date Range', f'{cluster.date_range_start} to {cluster.date_range_end}' if cluster.date_range_start else 'N/A'],
                ['Municipalities', ', '.join(cluster.municipalities) if cluster.municipalities else 'N/A'],
            ]

            detail_table = Table(detail_data, colWidths=[2*inch, 4*inch])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))

            elements.append(detail_table)
            elements.append(Spacer(1, 15))

        # Add disclaimer
        disclaimer = Paragraph(
            "<i>This report was generated using the AGNES (Agglomerative Nesting) clustering algorithm. "
            "Hotspots represent areas with statistically significant concentrations of traffic accidents. "
            "This information should be used for preventive measures and traffic safety planning.</i>",
            styles['Normal']
        )
        elements.append(Spacer(1, 30))
        elements.append(disclaimer)

        # Build PDF
        doc.build(elements)

        return filepath

    def generate_chart(self, queryset):
        """
        Generate severity distribution chart

        Args:
            queryset: AccidentCluster queryset

        Returns:
            BytesIO: Image buffer
        """
        severity_scores = [cluster.severity_score for cluster in queryset[:20]]
        locations = [cluster.primary_location[:20] for cluster in queryset[:20]]

        plt.figure(figsize=(10, 6))
        plt.barh(locations, severity_scores, color='#366092')
        plt.xlabel('Severity Score')
        plt.title('Top 20 Hotspots by Severity')
        plt.tight_layout()

        # Save to buffer
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=150)
        buffer.seek(0)
        plt.close()

        return buffer
